from __future__ import annotations

import json
import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_FILE = ROOT / "沪深北交易所《可持续发展报告指引（试行）》要求拆解.xlsx"
OUTPUT_FILE = ROOT / "static" / "guideline-data.js"

NUMBERED_TERM_RE = re.compile(r"^（[一二三四五六七八九十百千]+）\s*([^：:]+?)\s*[：:]\s*(.+)$")
PREV_TERM_RE = re.compile(r"^前款所称(.+?)是指(.+)$")
COLLECTIVE_TERM_RE = re.compile(r"^.+?以下统称(.+?)[。；;]?$")
GENERIC_TERM_RE = re.compile(r"^([^，。；;：:]{1,24}?)是指(.+)$")

TERM_ALIASES = {
    "可持续发展相关风险和机遇": ["风险和机遇"],
}

# ── 议题映射表（条号 → 议题）─────────────────────────────

ARTICLE_TOPICS = {
    # 格式：条号 → 议题名称
    21: "应对气候变化",
    22: "应对气候变化",
    23: "应对气候变化",
    24: "应对气候变化",
    25: "应对气候变化",
    26: "应对气候变化",
    27: "应对气候变化",
    28: "应对气候变化",
    30: "污染物排放",
    31: "废弃物处理",
    32: "生态系统和生物多样性保护",
    33: "环境合规管理",
    35: "能源利用",
    36: "水资源利用",
    37: "循环经济",
    39: "乡村振兴",
    40: "社会贡献",
    42: "创新驱动",
    43: "科技伦理",
    45: "供应链安全",
    46: "平等对待中小企业",
    47: "产品和服务安全与质量",
    48: "数据安全与客户隐私保护",
    50: "员工",
    52: "尽职调查",
    53: "利益相关方沟通",
    55: "反商业贿赂及反贪污",
    56: "反不正当竞争",
}


def extract_article_number(title: str) -> int | None:
    """从条文标题提取条号数字，如 '第二十一条' → 21"""
    if not title:
        return None
    # 尝试匹配"第XX条"，XX可能是中文数字或阿拉伯数字
    match = re.search(r"第([零一二三四五六七八九十百千\d]+)条", title)
    if not match:
        return None
    num_str = match.group(1)
    # 尝试转换中文数字或直接转为整数
    try:
        return int(num_str)
    except ValueError:
        return _parse_chinese_number(num_str)


def _parse_chinese_number(chinese_str: str) -> int | None:
    """将中文数字转换为阿拉伯数字"""
    chinese_num_map = {
        "零": 0, "一": 1, "二": 2, "三": 3, "四": 4, "五": 5,
        "六": 6, "七": 7, "八": 8, "九": 9, "十": 10, "百": 100, "千": 1000
    }
    if not chinese_str:
        return None
    try:
        # 简单实现：支持"第十五条"或"第一百二十条"等常见格式
        result = 0
        current = 0
        for char in chinese_str:
            if char in chinese_num_map:
                val = chinese_num_map[char]
                if val == 10:
                    current = max(1, current) * 10
                elif val == 100:
                    current = max(1, current) * 100
                elif val == 1000:
                    current = max(1, current) * 1000
                else:
                    current += val
            else:
                continue
        result = current
        return result if result > 0 else None
    except Exception:
        return None


def get_topic_for_article(article_title: str) -> str | None:
    """根据条文标题返回对应的议题"""
    article_num = extract_article_number(article_title)
    if article_num is None:
        return None
    return ARTICLE_TOPICS.get(article_num)


# ── HTML pre-rendering helpers ─────────────────────────────

def _escape(s: str) -> str:
    """Minimal HTML escape for text nodes and attribute values."""
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;").replace("'", "&#39;")


def build_term_pattern(glossary: list[dict]) -> re.Pattern | None:
    if not glossary:
        return None
    terms = sorted([g["term"] for g in glossary if g["term"]], key=len, reverse=True)
    if not terms:
        return None

    def term_to_pattern(t: str) -> str:
        p = re.escape(t)
        # "可持续发展" 作为独立词出现时才链接；在"可持续发展报告"中不链接
        if t == "可持续发展":
            p += r"(?!报告)"
        return p

    return re.compile("|".join(term_to_pattern(t) for t in terms))


def highlight_terms_to_html(text: str, pattern: re.Pattern | None) -> str:
    """Convert plain text to HTML with term-trigger buttons injected."""
    if not text:
        return ""
    if pattern is None:
        return _escape(text)
    parts: list[str] = []
    last = 0
    for m in pattern.finditer(text):
        parts.append(_escape(text[last : m.start()]))
        term = m.group()
        parts.append(
            f'<button type="button" class="term-trigger" data-term="{_escape(term)}">'
            f"{_escape(term)}</button>"
        )
        last = m.end()
    parts.append(_escape(text[last:]))
    return "".join(parts)


# ── Core helpers ─────────────────────────────────────────────

def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u3000", " ").replace("\xa0", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize_type(value: str) -> str:
    return value or "未分类"


def slugify(*parts: str) -> str:
    text = "-".join(part for part in parts if part)
    text = re.sub(r"[^\w\u4e00-\u9fff-]+", "-", text, flags=re.UNICODE)
    return re.sub(r"-+", "-", text).strip("-").lower()


def split_numbered_definitions(text: str) -> list[dict[str, str]]:
    matches = list(NUMBERED_TERM_RE.finditer(text))
    if not matches:
        return []

    items: list[dict[str, str]] = []
    for index, match in enumerate(matches):
        start = match.start()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        chunk = text[start:end].strip(" ；;。")
        term_match = NUMBERED_TERM_RE.match(chunk)
        if not term_match:
            continue
        term = clean_text(term_match.group(1))
        definition = clean_text(chunk)
        items.append({"term": term, "definition": definition})
    return items


def extract_terms(text: str) -> list[dict[str, str]]:
    text = clean_text(text)
    if not text:
        return []

    numbered_items = split_numbered_definitions(text)
    if numbered_items:
        return numbered_items

    matched = PREV_TERM_RE.match(text)
    if matched:
        return [{"term": clean_text(matched.group(1)), "definition": text}]

    matched = COLLECTIVE_TERM_RE.match(text)
    if matched:
        return [{"term": clean_text(matched.group(1)), "definition": text}]

    matched = GENERIC_TERM_RE.match(text)
    if matched:
        return [{"term": clean_text(matched.group(1)), "definition": text}]

    return []


def expand_term_aliases(glossary: OrderedDict[str, dict[str, str]]) -> OrderedDict[str, dict[str, str]]:
    expanded: OrderedDict[str, dict[str, str]] = OrderedDict(glossary)
    for term, entry in list(glossary.items()):
        for alias in TERM_ALIASES.get(term, []):
            if alias and alias not in expanded:
                expanded[alias] = {
                    "term": alias,
                    "definition": entry["definition"],
                }
    return expanded


def build_tree(rows: list[dict[str, str]]) -> list[dict[str, object]]:
    chapters: OrderedDict[str, dict[str, object]] = OrderedDict()

    for row in rows:
        chapter_name = row["chapter"] or "未分章"
        section_name = row["section"] or "未分节"
        article_name = row["article"] or "未分条"
        clause_name = row["clause"] or "未分款"

        chapter = chapters.setdefault(
            chapter_name,
            {
                "id": slugify(chapter_name),
                "title": chapter_name,
                "sections": OrderedDict(),
            },
        )
        section = chapter["sections"].setdefault(
            section_name,
            {
                "id": slugify(chapter_name, section_name),
                "title": section_name,
                "articles": OrderedDict(),
            },
        )
        article = section["articles"].setdefault(
            article_name,
            {
                "id": slugify(chapter_name, section_name, article_name),
                "title": article_name,
                "topic": get_topic_for_article(article_name) or "",
                "clauses": OrderedDict(),
            },
        )
        clause = article["clauses"].setdefault(
            clause_name,
            {
                "id": slugify(chapter_name, section_name, article_name, clause_name),
                "title": clause_name,
                "items": [],
            },
        )
        clause["items"].append(
            {
                "content": row["content"],
                "contentHtml": row.get("contentHtml", ""),
                "type": normalize_type(row["type"]),
                "disclosureLevel": row["disclosureLevel"],
            }
        )

    result: list[dict[str, object]] = []
    for chapter in chapters.values():
        sections: list[dict[str, object]] = []
        for section in chapter["sections"].values():
            articles: list[dict[str, object]] = []
            for article in section["articles"].values():
                clauses = list(article["clauses"].values())
                articles.append(
                    {
                        "id": article["id"],
                        "title": article["title"],
                        "topic": article["topic"],
                        "clauses": clauses,
                    }
                )
            sections.append(
                {
                    "id": section["id"],
                    "title": section["title"],
                    "articles": articles,
                }
            )
        result.append(
            {
                "id": chapter["id"],
                "title": chapter["title"],
                "sections": sections,
            }
        )
    return result


def load_rows() -> tuple[list[dict[str, str]], list[dict[str, str]], dict[str, int]]:
    workbook = load_workbook(SOURCE_FILE, read_only=True, data_only=True)
    sheet = workbook.active

    rows: list[dict[str, str]] = []
    glossary: OrderedDict[str, dict[str, str]] = OrderedDict()
    current_chapter = ""
    current_section = ""
    type_counter: dict[str, int] = {}

    for raw in sheet.iter_rows(min_row=3, values_only=True):
        raw_chapter = clean_text(raw[0])
        raw_section = clean_text(raw[1])
        article = clean_text(raw[2])
        clause = clean_text(raw[3])
        content = clean_text(raw[4])
        row_type = clean_text(raw[5])
        disclosure_level = clean_text(raw[6])

        if not any([raw_chapter, raw_section, article, clause, content, row_type, disclosure_level]):
            continue

        # 允许“章/节标题行”更新上下文，但不把空内容行写入数据。
        if raw_chapter:
            current_chapter = raw_chapter
        if raw_section:
            current_section = raw_section

        chapter = current_chapter
        section = current_section

        # 只有实际条文内容（条/款/内容/类型/披露级别任一非空）才生成记录。
        if not any([article, clause, content, row_type, disclosure_level]):
            continue

        type_name = normalize_type(row_type)
        type_counter[type_name] = type_counter.get(type_name, 0) + 1

        row = {
            "chapter": chapter,
            "section": section,
            "article": article,
            "clause": clause,
            "content": content,
            "type": type_name,
            "disclosureLevel": disclosure_level,
        }
        rows.append(row)

        if type_name == "释义":
            for item in extract_terms(content):
                term = item["term"]
                if term and term not in glossary:
                    glossary[term] = {
                        "term": term,
                        "definition": item["definition"],
                    }

    glossary = expand_term_aliases(glossary)
    return rows, list(glossary.values()), type_counter


def main() -> None:
    rows, glossary, type_counter = load_rows()

    # Pre-render term-highlighted HTML once at build time; avoids runtime regex
    pattern = build_term_pattern(glossary)
    for row in rows:
        row["contentHtml"] = highlight_terms_to_html(row["content"], pattern)
    for entry in glossary:
        entry["definitionHtml"] = highlight_terms_to_html(entry["definition"], pattern)

    payload = {
        "meta": {
            "title": "A股《可持续发展报告指引》浏览器（丸爸）",
            "sourceFile": SOURCE_FILE.name,
            "rowCount": len(rows),
            "termCount": len(glossary),
            "typeCount": type_counter,
        },
        "chapters": build_tree(rows),
        "glossary": glossary,
    }

    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    content = "window.GUIDELINE_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n"
    OUTPUT_FILE.write_text(content, encoding="utf-8")
    print(f"Wrote {OUTPUT_FILE.relative_to(ROOT)} with {len(rows)} rows and {len(glossary)} terms.")


if __name__ == "__main__":
    main()